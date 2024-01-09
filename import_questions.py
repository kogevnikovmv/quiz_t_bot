from openpyxl import load_workbook
from peewee import *

# excel файл
wb=load_workbook('C:\quiztbot\quest.xlsx')
ws=wb.active

# файл бд sqlite
db = SqliteDatabase('C:\quiztbot\questions_database.db')


# описание модели таблицы
class Question(Model):
    id=AutoField(column_name='id')
    question=TextField(column_name='question')
    answers=TextField(column_name='answers')
    right=TextField(column_name='right')
    picture=TextField(column_name='picture')
    question_type=TextField(column_name='question_type')


    class Meta:
        database = db
        table_name='Questions'


# функция добавления вопроса в бд
def add(question, answers, right, picture, question_type):
    row=Question(question=question.strip(), answers=answers, right=right, picture=picture, question_type=question_type)
    row.save()


# получение строки из xlsx и добавление в бд
for i in ws.iter_rows(min_col=1, max_col=7, min_row=1, max_row=ws.max_row, values_only=True):
    print(i[0], i[1], i[2], i[3], i[4], i[5], i[6])
    add(question=i[1], answers=i[2], right=i[3], picture=i[4], question_type=i[5])
        

print('all questions added')
db.close()